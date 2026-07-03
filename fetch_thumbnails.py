"""
One-time build step: fetch a small cover thumbnail for every book listed in
Book4You.xlsx directly from Microsoft Graph (OneDrive's own pre-rendered
thumbnails), with no need to download the actual PDFs.

Setup (one-time, in the Azure portal):
  1. https://portal.azure.com -> "Microsoft Entra ID" -> "App registrations" -> "New registration"
  2. Name: anything. Supported account types: "Personal Microsoft accounts only".
     Redirect URI: leave blank.
  3. After creation, copy the "Application (client) ID" from the Overview page.
  4. Go to "Authentication" -> turn on "Allow public client flows" = Yes -> Save.
  No client secret and no admin consent are needed.

Usage:
  python fetch_thumbnails.py <CLIENT_ID>

The first run opens a device-code sign-in (a URL + short code to enter in your
browser). A cached token is stored in token_cache.bin so you won't need to sign
in again on later runs.
"""

import json
import sys
import time
from pathlib import Path

import msal
import openpyxl
import requests

AUTHORITY = "https://login.microsoftonline.com/consumers"
SCOPES = ["Files.Read"]
TOKEN_CACHE_PATH = Path(__file__).parent / "token_cache.bin"
THUMBNAILS_DIR = Path(__file__).parent / "thumbnails"
DATA_OUT = Path(__file__).parent / "books_data.json"
WORKBOOK = Path(__file__).parent / "Book4You.xlsx"


def get_access_token(client_id: str) -> str:
    cache = msal.SerializableTokenCache()
    if TOKEN_CACHE_PATH.exists():
        cache.deserialize(TOKEN_CACHE_PATH.read_text())

    app = msal.PublicClientApplication(client_id, authority=AUTHORITY, token_cache=cache)

    accounts = app.get_accounts()
    result = None
    if accounts:
        result = app.acquire_token_silent(SCOPES, account=accounts[0])

    if not result:
        flow = app.initiate_device_flow(scopes=SCOPES)
        if "user_code" not in flow:
            raise RuntimeError(f"Failed to create device flow: {flow}")
        print(flow["message"])
        result = app.acquire_token_by_device_flow(flow)

    if cache.has_state_changed:
        TOKEN_CACHE_PATH.write_text(cache.serialize())

    if "access_token" not in result:
        raise RuntimeError(f"Auth failed: {result.get('error_description')}")

    return result["access_token"]


def read_books() -> list[dict]:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["Books4You"]
    books = []
    for r in range(8, ws.max_row + 1):
        item_id = ws.cell(row=r, column=1).value
        initial = ws.cell(row=r, column=2).value
        author_en = ws.cell(row=r, column=3).value
        alphabet = ws.cell(row=r, column=4).value
        author_mm = ws.cell(row=r, column=5).value
        title = ws.cell(row=r, column=6).value
        weburl = ws.cell(row=r, column=8).value
        filesize_mb = ws.cell(row=r, column=9).value
        if not item_id or not title:
            continue
        books.append(
            {
                "id": item_id,
                "initial": initial,
                "author_en": author_en,
                "alphabet": alphabet,
                "author_mm": author_mm,
                "title": str(title).rsplit(".pdf", 1)[0],
                "weburl": weburl,
                "filesize_mb": filesize_mb,
            }
        )
    return books


def _get_with_retries(url: str, headers: dict | None = None, attempts: int = 4):
    for attempt in range(attempts):
        try:
            return requests.get(url, headers=headers, timeout=30)
        except requests.exceptions.RequestException:
            if attempt == attempts - 1:
                raise
            time.sleep(2**attempt)


def fetch_thumbnail(token: str, item_id: str) -> bytes | None:
    headers = {"Authorization": f"Bearer {token}"}

    urls_to_try = [f"https://graph.microsoft.com/v1.0/me/drive/items/{item_id}/thumbnails"]
    if "!" in item_id:
        drive_id, real_id = item_id.split("!", 1)
        urls_to_try.append(
            f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{real_id}/thumbnails"
        )

    for url in urls_to_try:
        resp = _get_with_retries(url, headers=headers)
        if resp.status_code != 200:
            continue
        values = resp.json().get("value", [])
        if not values:
            continue
        thumb = values[0].get("medium") or values[0].get("small") or values[0].get("large")
        if not thumb:
            continue
        img_resp = _get_with_retries(thumb["url"])
        if img_resp.status_code == 200:
            return img_resp.content

    return None


def main():
    if len(sys.argv) < 2:
        print("Usage: python fetch_thumbnails.py <CLIENT_ID>")
        sys.exit(1)

    client_id = sys.argv[1]
    token = get_access_token(client_id)

    books = read_books()
    THUMBNAILS_DIR.mkdir(exist_ok=True)

    ok, missing = 0, 0
    for i, book in enumerate(books, 1):
        safe_id = book["id"].replace("!", "_")
        thumb_path = THUMBNAILS_DIR / f"{safe_id}.jpg"
        book["thumbnail"] = f"thumbnails/{safe_id}.jpg"

        if thumb_path.exists():
            ok += 1
        else:
            try:
                data = fetch_thumbnail(token, book["id"])
            except requests.exceptions.RequestException as e:
                print(f"  network error on {book['title']!r}: {e}")
                data = None
            if data:
                thumb_path.write_bytes(data)
                ok += 1
            else:
                book["thumbnail"] = None
                missing += 1

        if i % 25 == 0 or i == len(books):
            print(f"{i}/{len(books)} processed (ok={ok}, missing={missing})")
            DATA_OUT.write_text(json.dumps(books, ensure_ascii=False, indent=2), encoding="utf-8")

    DATA_OUT.write_text(json.dumps(books, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Done. {ok} thumbnails saved, {missing} missing. Data written to {DATA_OUT}")


if __name__ == "__main__":
    main()
